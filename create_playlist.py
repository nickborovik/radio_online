# -*- coding: utf-8 -*-
import sys
import smtplib
from pathlib import Path
from configparser import ConfigParser
from openpyxl import load_workbook
from mutagen import MutagenError
from mutagen.mp3 import MP3
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, date, time, timedelta

# STATIC SETTINGS

# DATES

CUR_DAY = datetime.strptime(sys.argv[1], '%d.%m.%Y').date()
TM_DAY = CUR_DAY + timedelta(days=1)
AT_DAY = TM_DAY + timedelta(days=1)

MONTH = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
         'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']

# MAIN DIRS

ROOT_DIR = Path('D:/')
BASE_DIR = ROOT_DIR / 'INTERNET RADIO'
ARCH_DIR = BASE_DIR / 'Archive_2018'
CONF_DIR = BASE_DIR / 'Playlist_auto_generator'
DO_15_DIR = ARCH_DIR / 'domashniy ochag 15 min'

KIEV_ST_DIR_TODAY = BASE_DIR / 'Kievskaya Studia' / f'!{CUR_DAY.strftime("%m %Y")}'
KIEV_ST_DIR_TOMOR = BASE_DIR / 'Kievskaya Studia' / f'!{TM_DAY.strftime("%m %Y")}'

KHAR_ST_DIR_TODAY = BASE_DIR / 'KharkovTWR' / '1 SREDNIE VOLNI ONLINE' / f'{CUR_DAY.strftime("%m-%Y")}'
KHAR_ST_DIR_TOMOR = BASE_DIR / 'KharkovTWR' / '1 SREDNIE VOLNI ONLINE' / f'{TM_DAY.strftime("%m-%Y")}'

# EXCEL settings

EXCEL_FILE_NAME = f'{TM_DAY.strftime("%m-%Y")} Расписание онлайн вещания ({MONTH[TM_DAY.month - 1]}).xlsx'
EXCEL_FILE_PATH = BASE_DIR / EXCEL_FILE_NAME
EXCEL_PAGE_NAME = f'{TM_DAY.strftime("%#d.%m")}'

# PLAYLIST settings

PLAYLIST_DIR = ROOT_DIR / 'Playlist Radioboss'
CUR_PLAYLIST_NAME = f'playlist_for_{TM_DAY.strftime("%d%m%Y")}.m3u8'
TM_PLAYLIST_NAME = f'playlist_for_{AT_DAY.strftime("%d%m%Y")}.m3u8'
CUR_PLAYLIST_PATH = PLAYLIST_DIR / CUR_PLAYLIST_NAME
TM_PLAYLIST_PATH = PLAYLIST_DIR / TM_PLAYLIST_NAME

# MISTAKE REPORT settings

PL_NOT_DONE_SUBJECT = f"Плейлист {CUR_PLAYLIST_NAME} для онлайн радио ТМР не был создан"
PL_DURATION_SUBJECT = f"Длительность плейлиста {CUR_PLAYLIST_NAME} свыше 24 часов и 10 минут"
PL_TRACKS_TIME = f"Серьезное отклонение передач по времени в плейлисте {CUR_PLAYLIST_NAME}"
NEW_LIVE_PROGRAM_MESSAGE = "Обнаружена новая программа в расписании!\n" \
                           "Нужно сообщить программисту о добавлении предачи:\n" \
                           "{}"

# MP3 FILES

MUZBLOCKS = {
    535: '11 Kharkov time 9 min-a.mp3',
    540: '12 Kharkov time 9 min-a.mp3',
    575: '14 Kharkov time 10 min-a.mp3',
    585: '13 Kharkov time 10 min-a.mp3',
    595: '15 Kharkov time 10 min-a.mp3',
    650: '09 Kharkov time 11 min-a.mp3',
    655: '04 Kharkov time 11 min-a.mp3',
    658: '10 Kharkov time 11 min-a.mp3',
    660: '03 Kharkov time 11 min-a.mp3',
    662: '01 Kharkov time 11 min-a.mp3',
    664: '06 Kharkov time 11 min-a.mp3',
    666: '05 Kharkov time 11 min-a.mp3',
    668: '02 Kharkov time 11 min-a.mp3',
    669: '08 Kharkov time 11 min-a.mp3',
    670: '07 Kharkov time 11 min-a.mp3',
    725: 'muzblok_01_time_12.15.mp3',
    740: 'muzblok_18_time_12.40.mp3',
    750: 'muzblok_12_time_12.40.mp3',
    770: 'muzblok_05_time_13.20.mp3',
    810: 'muzblok_03_time_13.42.mp3',
    822: 'muzblok_08_time_13.55.mp3',
    830: 'muzblok_15_time_13.58.mp3',
    835: 'muzblok_11_time_14.02.mp3',
    850: 'muzblok_24_time_14.16.mp3',
    855: 'muzblok_06_time_14.19.mp3',
    856: 'muzblok_09_time_14.20.mp3',
    857: 'muzblok_07_time_14.21.mp3',
    866: 'muzblok_13_time_14.27.mp3',
    872: 'muzblok_10_time_14.33.mp3',
    880: 'muzblok_14_time_14.41.mp3',
    882: 'muzblok_02_time_14.43.mp3',
    885: 'muzblok_17_time_14.46.mp3',
    888: 'muzblok_16_time_14.49.mp3',
    894: 'muzblok_23_time_14.55.mp3',
    894: 'muzblok_19_time_14.55.mp3',
    896: 'muzblok_04_time_14.57.mp3',
    898: 'muzblok_20_time_14.59.mp3',
    932: 'muzblok_22_time_15.33.mp3',
    942: 'muzblok_21_time_15.43.mp3',
    975: 'muzblok_26_time_16.16.mp3',
    977: 'muzblok_25_time_16.18.mp3',
}

LIVE_FILES = {
    '900 секунд доброты': ['RUS_KIND_{}.mp3', 'Kharkov'],
    'БА': ['RUS_BST_{}.mp3', 'Kiev'],
    'Библейские искатели': ['RUS_TSK_{}.mp3', 'Kiev'],
    'Блокнот миссионера': ['RUS_MC_{}.mp3', 'Kiev'],
    'Вивчаємо Біблію разом': ['UKR_SBT_{}.mp3', 'Kharkov'],
    'ВЦП': ['UKR_PRC_{}.mp3', 'Kiev'],
    'Герои': ['RUS_CA_{}.mp3', 'Kharkov'],
    'Голос друга': ['BEL_VFR_{}.mp3', 'Kiev'],
    'Джерельце': ['UKR_TLS_{}.mp3', 'Kiev'],
    'ЖКОЕ': ['RUS_LAI_{}.mp3', 'Kharkov'],
    'ЖН': ['UKR_HOPE_{}.mp3', 'Kiev'],
    'Калейдоскоп': ['UKR_KAL_{}.mp3', 'Kiev'],
    'МН': ['RUS_BOH_{}.mp3', 'Kiev'],
    'Ответственность': ['RUS_SPOT_{}.mp3', 'Kiev'],
    'Погляд ': ['UKR_TOV_{}.mp3', 'Kiev'],
    'Свет жизни': ['RUS_IFL_{}.mp3', 'Kiev'],
    'Серебро': ['RUS_SIL_{}.mp3', 'Kiev'],
    'Слово на сегодня': ['RUS_TWT_{}.mp3', 'Kiev'],
    'Стежинка': ['UKR_TLP_{}.mp3', 'Kiev'],
    'Суламита': ['RUS_SUL_{}.mp3', 'Kiev'],
    'Табор': ['RUS_RCMO_{}.mp3', 'Kharkov'],
    'Тихие воды': ['RUS_SWA_{}.mp3', 'Kiev'],
    'Хлеб жизни': ['RUS_BLR_{}.mp3', 'Kiev'],
    'Шалом': ['RUS_SHA_{}.mp3', 'Kharkov'],
    'Шанс // ГВЛ': ['UKR_MAE_{}.mp3', 'Kiev'],
}


# MAIN CODE

def send_email_report(subject, body_text):
    """Отправка письма в случае ошибки создания плейлиста"""
    config_path = CONF_DIR / "email.ini"

    if config_path.exists():
        cfg = ConfigParser()
        cfg.read(config_path)
    else:
        print("Файл конфигурации email.ini не найден!")
        raise SystemExit

    host = cfg.get("smtp", "host")
    from_addr = cfg.get("smtp", "from_addr")
    password = cfg.get("smtp", "password")
    to_emails = cfg.get("smtp", "to_emails").split(',')

    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = ', '.join(to_emails)
    msg['Subject'] = Header(subject)

    msg.attach(MIMEText(body_text, 'plain', 'cp1251'))

    server = smtplib.SMTP_SSL(host, 465)
    server.ehlo()
    server.login(user=from_addr, password=password)
    server.sendmail(from_addr, to_emails, msg.as_string())
    server.quit()


def get_excel_sheet(file_name, excel_page_name):
    """Возвращает лист с книги Excel"""
    if file_name.exists():
        workbook = load_workbook(file_name, data_only=True)
        sheet = workbook[excel_page_name]
        return sheet

    email_text = f"Excel файл \n" \
                 f"{file_name.absolute()}\n" \
                 f"Не найден, проверьте файл в папке 'INTERNET RADIO'"
    send_email_report(PL_NOT_DONE_SUBJECT, email_text)
    print(email_text)
    raise SystemExit


def get_muzblock(row, tracks_time_total):
    """Выбрать самый подходящий музблок"""
    h, m, s = (23, 59, 59) if row[2] == time(0, 0) else (row[2].hour, row[2].minute, 0)
    track_end_time = timedelta(hours=h, minutes=m, seconds=s)
    tracks_time = timedelta(seconds=tracks_time_total)
    file_duration = int((track_end_time - tracks_time).total_seconds())
    track = min(MUZBLOCKS, key=lambda x: abs(x - file_duration))
    return MUZBLOCKS[track]


def get_excel_data(row, tracks_time_total):
    if row[5] == 'муз.блок':
        """Получаем музблок"""
        file_title = 'Muzblock'
        file_name = get_muzblock(row, tracks_time_total)
        file_path = ARCH_DIR / file_name

    elif row[5] == 'ГОДИНА БОЖОГО СЛОВА':
        """Конкретный случай для передачи Година Божого Слова"""
        file_num = row[4]
        file_title = 'Online radio blok'
        file_name = f'{file_title} {file_num}.mp3'
        file_path = ARCH_DIR / file_name

    elif row[5] == 'ДО (15)':
        """Конкретный случай для передачи Домашний очаг 15 минут"""
        file_num = row[4]
        file_title = row[3]
        file_name = f'{file_title} {file_num}.mp3'
        file_path = DO_15_DIR / file_name

    elif row[5] == 'Про сімейні цінності':
        file_num = row[4]
        file_title = 'Pro simeyni zinnosti'
        file_name = f'{file_title} {file_num}.mp3'
        file_path = ARCH_DIR / file_name

    elif row[3] == 'Live':
        file_title = row[3]
        file_num = row[4]
        file_name = f'{file_title} {file_num}.mp3'
        file_path = ARCH_DIR / file_name
        if not file_path.exists():
            file_num = 0
            file_name = f'{file_title} {file_num}.mp3'
            file_path = ARCH_DIR / file_name

    elif time(10, 0) > row[1] >= time(8, 30):
        """Повтор за вчера"""
        if row[5] not in LIVE_FILES:
            error = NEW_LIVE_PROGRAM_MESSAGE.format(row[5])
            print(error)
            send_email_report(PL_NOT_DONE_SUBJECT, error)
            raise SystemExit
        date = CUR_DAY.strftime('%Y%m%d')
        file_title = LIVE_FILES[row[5]][0].format(date)
        file_name = LIVE_FILES[row[5]][0].format(date)
        file_dir = KIEV_ST_DIR_TODAY if LIVE_FILES[row[5]][1] == 'Kiev' else KHAR_ST_DIR_TODAY
        file_path = file_dir / file_name

    elif time(22, 0) > row[1] >= time(20, 30):
        """Прямой эфир"""
        if row[5] not in LIVE_FILES:
            error = NEW_LIVE_PROGRAM_MESSAGE.format(row[5])
            print(error)
            send_email_report(PL_NOT_DONE_SUBJECT, error)
            raise SystemExit
        date = TM_DAY.strftime('%Y%m%d')
        file_title = LIVE_FILES[row[5]][0].format(date)
        file_name = LIVE_FILES[row[5]][0].format(date)
        file_dir = KIEV_ST_DIR_TOMOR if LIVE_FILES[row[5]][1] == 'Kiev' else KHAR_ST_DIR_TOMOR
        file_path = file_dir / file_name

    else:
        """Все остальные случаи, где файл из папки Archive_2018"""
        file_num = str(row[4])
        if 'Лекция' in file_num:
            file_num = file_num.replace('Лекция', 'L')
        if 'М.В.' in file_num:
            file_num = file_num.replace('М.В.', 'M')
        if 'А.М.' in file_num:
            file_num = file_num.replace('А.М.', 'A')
        file_title = row[3]
        file_name = f'{file_title} {file_num}.mp3'.replace('  ', ' ')
        file_path = ARCH_DIR / file_name

    return file_title, file_path


def get_file_duration(file_path, list_duration, time_start):
    """Возвращает длинну MP3 трека в секундах"""
    if file_path.exists():
        error = ''
        try:
            mp3 = MP3(file_path)
            mp3_duration = int(mp3.info.length)
            track_duration = timedelta(seconds=mp3_duration)
            length = (datetime.combine(date.today(), list_duration) - track_duration).time()
            if time(0, 5) < length < time(23, 55):
                error = f'Трек: {file_path}\n' \
                        f'Время в списке: {list_duration}\n' \
                        f'Время трека: {track_duration}\n'
            return int(mp3.info.length), error
        except MutagenError:
            email_text = f"MP3 файл\n" \
                         f"{file_path.absolute()}\n" \
                         f"Поврежден и (или) не может быть открыт\n" \
                         f"Проверьте состояние файла"
            send_email_report(PL_NOT_DONE_SUBJECT, email_text)
            print(email_text)
            raise MutagenError
        except Exception:
            email_text = f"Во время считывания файла\n---\n{file_path.absolute()}\n---\n" \
                         f"Сборщик плейлистов завершил работу\n{Exception}"
            send_email_report(PL_NOT_DONE_SUBJECT, email_text)
            raise Exception

    email_text = f"MP3 файл\n" \
                 f"{file_path.absolute()}\n" \
                 f"Не найден, проверьте наличие файла в папке\n" \
                 f"Страница в файле EXCEL: {EXCEL_PAGE_NAME}\n" \
                 f"Время начала передачи: {time_start}"
    print(email_text)
    send_email_report(PL_NOT_DONE_SUBJECT, email_text)
    raise SystemExit


def write_playlist(playlist_path, playlist_data):
    """Записать плейлист в файл"""
    with open(playlist_path, 'w') as write_file:
        write_file.writelines(playlist_data)
    print(f'Плейлист на {TM_DAY.strftime("%d.%m.%Y")} готов и находится в папке\n'
          f'{PLAYLIST_DIR.absolute()}')


def main():
    playlist_data = ['#EXTM3U\n']
    sheet = get_excel_sheet(EXCEL_FILE_PATH, EXCEL_PAGE_NAME)
    tracks_time_total = 0
    errors = []

    for row in sheet.iter_rows(min_row=4, max_col=7, values_only=True):
        if not any(row[3:6]):
            continue
        file_name, file_path = get_excel_data(row, tracks_time_total)
        file_duration, error = get_file_duration(file_path, row[6], row[1])
        if error and row[5] != 'муз.блок':
            errors.append(error)
        tracks_time_total += file_duration
        playlist_data.append(f'#EXTINF:{file_duration},{file_name}\n{file_path}\n')

    playlist_data.append(f'load {TM_PLAYLIST_PATH}.command')
    write_playlist(CUR_PLAYLIST_PATH, playlist_data)

    """Проверить длительность плейлиста"""
    if 84600 > tracks_time_total > 87000:
        pl_time = str(timedelta(seconds=tracks_time_total))
        email_text = f"Плейлист\n" \
                     f"{CUR_PLAYLIST_NAME}\n" \
                     f"собран, но его продолжительность {pl_time}"
        print(email_text)
        send_email_report(PL_DURATION_SUBJECT, email_text)

    if errors:
        email_text = 'Следующие треки не совпадают по времени\n' + '\n'.join(errors)
        print(email_text)
        send_email_report(PL_TRACKS_TIME, email_text)


if __name__ == '__main__':
    main()
